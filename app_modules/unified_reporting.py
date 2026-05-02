"""
Unified Reporting Module
==========================
Single source of truth for all report generation with rich Excel exports.
Supports multiple sheets, formatting, charts, and comprehensive analytics.
"""

from __future__ import annotations

import io
from datetime import datetime, date
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage


# ==========================
#  DATA CLASSES
# ==========================
@dataclass
class ReportSection:
    """Defines a section in the report."""
    title: str
    dataframe: pd.DataFrame
    description: str = ""
    chart_type: Optional[str] = None  # 'bar', 'pie', 'line'
    chart_column: Optional[str] = None  # Column for chart data
    chart_figure: Optional[Any] = None  # Plotly Figure


@dataclass
class ReportMetadata:
    """Metadata for the report."""
    title: str = "Analytics Report"
    generated_by: str = "Automation Hub Pro"
    generated_at: datetime = field(default_factory=datetime.now)
    date_range: Optional[Tuple[date, date]] = None
    filters_applied: List[str] = field(default_factory=list)
    total_records: int = 0


# ==========================
#  STYLES CONFIGURATION
# ==========================
class ReportStyles:
    """Centralized styling for Excel reports."""
    
    # Colors
    PRIMARY = "1d4ed8"
    PRIMARY_LIGHT = "dbeafe"
    SUCCESS = "10b981"
    SUCCESS_LIGHT = "d1fae5"
    WARNING = "f59e0b"
    WARNING_LIGHT = "fef3c7"
    DANGER = "ef4444"
    DANGER_LIGHT = "fee2e2"
    DARK = "1e293b"
    GRAY = "64748b"
    LIGHT_GRAY = "f1f5f9"
    
    # Fonts
    HEADER_FONT = Font(name='Calibri', size=14, bold=True, color=PRIMARY)
    TITLE_FONT = Font(name='Calibri', size=12, bold=True, color=DARK)
    SUBTITLE_FONT = Font(name='Calibri', size=11, bold=True, color=GRAY)
    NORMAL_FONT = Font(name='Calibri', size=10)
    BOLD_FONT = Font(name='Calibri', size=10, bold=True)
    
    # Fills
    HEADER_FILL = PatternFill(start_color=PRIMARY_LIGHT, end_color=PRIMARY_LIGHT, fill_type="solid")
    SUCCESS_FILL = PatternFill(start_color=SUCCESS_LIGHT, end_color=SUCCESS_LIGHT, fill_type="solid")
    WARNING_FILL = PatternFill(start_color=WARNING_LIGHT, end_color=WARNING_LIGHT, fill_type="solid")
    DANGER_FILL = PatternFill(start_color=DANGER_LIGHT, end_color=DANGER_LIGHT, fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    
    # Borders
    THIN_BORDER = Border(
        left=Side(style='thin', color='e2e8f0'),
        right=Side(style='thin', color='e2e8f0'),
        top=Side(style='thin', color='e2e8f0'),
        bottom=Side(style='thin', color='e2e8f0')
    )
    
    # Alignments
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
    RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')


# ==========================
#  UNIFIED REPORT GENERATOR
# ==========================
class UnifiedReportGenerator:
    """Single source of truth for all report generation."""
    
    def __init__(self, metadata: Optional[ReportMetadata] = None):
        self.metadata = metadata or ReportMetadata()
        self.styles = ReportStyles()
        self.sections: List[ReportSection] = []
    
    def add_section(self, section: ReportSection) -> 'UnifiedReportGenerator':
        """Add a data section to the report."""
        self.sections.append(section)
        self.metadata.total_records += len(section.dataframe)
        return self
    
    def generate_excel(self) -> bytes:
        """Generate rich Excel report with all sections."""
        wb = Workbook()
        
        # Remove default sheet and create summary
        wb.remove(wb.active)
        
        # Create Summary Sheet
        self._create_summary_sheet(wb)
        
        # Create Data Sheets
        for section in self.sections:
            self._create_data_sheet(wb, section)
        
        # Create Analytics Sheet (if charts enabled)
        if any(s.chart_type for s in self.sections):
            self._create_analytics_sheet(wb)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    
    def _create_summary_sheet(self, wb: Workbook):
        """Create executive summary sheet."""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = self.metadata.title
        ws['A1'].font = Font(name='Calibri', size=18, bold=True, color=self.styles.PRIMARY)
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = self.styles.CENTER_ALIGN
        
        # Metadata section
        row = 3
        ws[f'A{row}'] = "Report Information"
        ws[f'A{row}'].font = self.styles.TITLE_FONT
        ws[f'A{row}'].fill = self.styles.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        info_data = [
            ("Generated By:", self.metadata.generated_by),
            ("Generated At:", self.metadata.generated_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Records:", f"{self.metadata.total_records:,}"),
        ]
        
        if self.metadata.date_range:
            info_data.append(("Date Range:", f"{self.metadata.date_range[0]} to {self.metadata.date_range[1]}"))
        
        if self.metadata.filters_applied:
            info_data.append(("Filters Applied:", ", ".join(self.metadata.filters_applied)))
        
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.styles.BOLD_FONT
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = self.styles.NORMAL_FONT
            row += 1
        
        # Sections summary
        row += 2
        ws[f'A{row}'] = "Report Sections"
        ws[f'A{row}'].font = self.styles.TITLE_FONT
        ws[f'A{row}'].fill = self.styles.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        headers = ["Section", "Records", "Description", "Has Chart"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.BOLD_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER_ALIGN
            cell.border = self.styles.THIN_BORDER
        
        for section in self.sections:
            row += 1
            ws.cell(row=row, column=1, value=section.title).border = self.styles.THIN_BORDER
            ws.cell(row=row, column=2, value=len(section.dataframe)).border = self.styles.THIN_BORDER
            ws.cell(row=row, column=3, value=section.description).border = self.styles.THIN_BORDER
            ws.cell(row=row, column=4, value="Yes" if section.chart_type else "No").border = self.styles.THIN_BORDER
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
    
    def _create_data_sheet(self, wb: Workbook, section: ReportSection):
        """Create a formatted data sheet."""
        # Sanitize sheet name
        sheet_name = section.title[:31]  # Excel limit
        ws = wb.create_sheet(sheet_name)
        
        # Sheet title
        ws['A1'] = section.title
        ws['A1'].font = self.styles.HEADER_FONT
        ws.merge_cells(f'A1:{get_column_letter(len(section.dataframe.columns))}1')
        
        if section.description:
            ws['A2'] = section.description
            ws['A2'].font = self.styles.SUBTITLE_FONT
            ws.merge_cells(f'A2:{get_column_letter(len(section.dataframe.columns))}2')
            start_row = 4
        else:
            start_row = 3
        
        # Headers
        for col_idx, col_name in enumerate(section.dataframe.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.font = self.styles.BOLD_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER_ALIGN
            cell.border = self.styles.THIN_BORDER
        
        # Data rows with alternating colors
        for row_idx, row_data in enumerate(section.dataframe.values, start=start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles.NORMAL_FONT
                cell.border = self.styles.THIN_BORDER
                
                # Alternating row colors
                if (row_idx - start_row) % 2 == 0:
                    cell.fill = self.styles.ALT_ROW_FILL
                
                # Number formatting
                if isinstance(value, (int, float)):
                    cell.alignment = self.styles.RIGHT_ALIGN
                    if 'price' in str(section.dataframe.columns[col_idx-1]).lower() or \
                       'amount' in str(section.dataframe.columns[col_idx-1]).lower() or \
                       'revenue' in str(section.dataframe.columns[col_idx-1]).lower():
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = self.styles.LEFT_ALIGN
        
        # Add chart if specified
        if section.chart_figure:
            self._add_plotly_chart(ws, section, start_row)
        elif section.chart_type and section.chart_column:
            self._add_chart(ws, section, start_row, len(section.dataframe) + start_row)
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
    
    def _add_plotly_chart(self, ws, section: ReportSection, start_row: int):
        """Render a Plotly figure as a static PNG and insert it into the Excel sheet."""
        try:
            # Requires 'kaleido' package to be installed
            img_bytes = section.chart_figure.to_image(format="png", width=800, height=500)
            img = OpenpyxlImage(io.BytesIO(img_bytes))
            col_idx = len(section.dataframe.columns) + 2
            cell_anchor = f"{get_column_letter(col_idx)}3"
            ws.add_image(img, cell_anchor)
        except Exception as e:
            # Fallback or silent ignore if kaleido is missing/fails
            import logging
            logging.error(f"Failed to render Plotly image to Excel: {e}")
            col_idx = len(section.dataframe.columns) + 2
            ws[f"{get_column_letter(col_idx)}3"] = f"Chart generation failed. (Ensure 'kaleido' is installed via pip): {e}"

    def _add_chart(self, ws, section: ReportSection, start_row: int, end_row: int):
        """Add chart to worksheet."""
        if section.chart_type == 'bar':
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = f"{section.title} Analysis"
            
            # Find data column
            if section.chart_column in section.dataframe.columns:
                col_idx = list(section.dataframe.columns).index(section.chart_column) + 1
                data = Reference(ws, min_col=col_idx, min_row=start_row, max_row=end_row)
                cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=end_row)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                ws.add_chart(chart, f"{get_column_letter(col_idx + 2)}3")
        
        elif section.chart_type == 'pie':
            chart = PieChart()
            chart.title = f"{section.title} Distribution"
            
            if section.chart_column in section.dataframe.columns:
                col_idx = list(section.dataframe.columns).index(section.chart_column) + 1
                labels = Reference(ws, min_col=1, min_row=start_row + 1, max_row=min(end_row, start_row + 10))
                data = Reference(ws, min_col=col_idx, min_row=start_row, max_row=min(end_row, start_row + 10))
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(labels)
                ws.add_chart(chart, f"{get_column_letter(col_idx + 2)}3")
    
    def _create_analytics_sheet(self, wb: Workbook):
        """Create analytics overview sheet with KPIs."""
        ws = wb.create_sheet("Analytics")
        
        ws['A1'] = "Analytics Overview"
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=self.styles.PRIMARY)
        ws.merge_cells('A1:D1')
        
        row = 3
        for section in self.sections:
            if len(section.dataframe) > 0:
                ws[f'A{row}'] = f"{section.title} Metrics"
                ws[f'A{row}'].font = self.styles.TITLE_FONT
                ws[f'A{row}'].fill = self.styles.HEADER_FILL
                ws.merge_cells(f'A{row}:D{row}')
                row += 1
                
                # Calculate basic metrics
                numeric_cols = section.dataframe.select_dtypes(include=['number']).columns
                for col in numeric_cols[:3]:  # Limit to first 3 numeric columns
                    stats = section.dataframe[col].describe()
                    metrics = [
                        (f"{col} - Total", f"{stats.get('count', 0):,.0f}"),
                        (f"{col} - Mean", f"{stats.get('mean', 0):,.2f}"),
                        (f"{col} - Max", f"{stats.get('max', 0):,.2f}"),
                    ]
                    for metric_name, metric_value in metrics:
                        ws[f'A{row}'] = metric_name
                        ws[f'A{row}'].font = self.styles.BOLD_FONT
                        ws[f'B{row}'] = metric_value
                        ws[f'B{row}'].font = self.styles.NORMAL_FONT
                        row += 1
                
                row += 1
        
        self._adjust_column_widths(ws)
    
    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width


# ==========================
#  STREAMLIT UI COMPONENTS
# ==========================
def render_unified_export_section(
    sections: List[ReportSection],
    metadata: Optional[ReportMetadata] = None,
    filename_prefix: str = "report"
) -> None:
    """
    Render unified export section in Streamlit UI.
    
    Args:
        sections: List of ReportSection to include
        metadata: Report metadata
        filename_prefix: Prefix for download filename
    """
    st.divider()
    st.subheader("📊 Export & Reports")
    
    if not sections:
        st.info("No data available for export.")
        return
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("**📄 Quick Exports**")
        
        # Individual CSV downloads
        for section in sections:
            if not section.dataframe.empty:
                csv = section.dataframe.to_csv(index=False).encode('utf-8-sig')
                safe_name = section.title.replace(" ", "_").lower()[:30]
                st.download_button(
                    label=f"⬇️ {section.title} (CSV)",
                    data=csv,
                    file_name=f"{filename_prefix}_{safe_name}.csv",
                    mime="text/csv",
                    use_container_width=True,
                    key=f"csv_{safe_name}_{id(section)}"
                )
    
    with col2:
        st.markdown("**📑 Excel Report**")
        
        # Rich Excel report
        generator = UnifiedReportGenerator(metadata=metadata)
        for section in sections:
            generator.add_section(section)
        
        try:
            excel_bytes = generator.generate_excel()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(
                label="⬇️ Download Full Report (Excel)",
                data=excel_bytes,
                file_name=f"{filename_prefix}_report_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
                key=f"excel_full_{timestamp}"
            )
            
            # Show report preview
            with st.expander("📋 Report Preview", expanded=False):
                st.markdown(f"**Report:** {metadata.title if metadata else 'Analytics Report'}")
                st.markdown(f"**Sections:** {len(sections)}")
                st.markdown(f"**Total Records:** {sum(len(s.dataframe) for s in sections):,}")
                for section in sections:
                    st.markdown(f"- {section.title}: {len(section.dataframe):,} records")
                    
        except Exception as e:
            st.error(f"Failed to generate Excel report: {str(e)}")
    
    with col3:
        st.markdown("**📈 Statistics**")
        
        total_records = sum(len(s.dataframe) for s in sections)
        st.metric("Total Records", f"{total_records:,}")
        
        # Per-section counts
        for section in sections:
            st.caption(f"{section.title}: {len(section.dataframe):,}")


def create_report_section(
    title: str,
    df: pd.DataFrame,
    description: str = "",
    chart_type: Optional[str] = None,
    chart_column: Optional[str] = None,
    chart_figure: Optional[Any] = None
) -> ReportSection:
    """Helper to create a ReportSection from a DataFrame."""
    return ReportSection(
        title=title,
        dataframe=df.copy(),
        description=description,
        chart_type=chart_type,
        chart_column=chart_column,
        chart_figure=chart_figure
    )


# ==========================
#  LEGACY COMPATIBILITY
# ==========================
def to_excel_bytes_enhanced(
    df: pd.DataFrame, 
    sheet_name: str = "Sheet1",
    add_chart: bool = False,
    chart_column: Optional[str] = None
) -> bytes:
    """
    Enhanced Excel export (backward compatible with to_excel_bytes).
    
    Args:
        df: DataFrame to export
        sheet_name: Name of the sheet
        add_chart: Whether to add a chart
        chart_column: Column to use for chart data
    """
    section = ReportSection(
        title=sheet_name,
        dataframe=df,
        chart_type='bar' if add_chart and chart_column else None,
        chart_column=chart_column
    )
    
    metadata = ReportMetadata(
        title=f"{sheet_name} Report",
        total_records=len(df)
    )
    
    generator = UnifiedReportGenerator(metadata=metadata)
    generator.add_section(section)
    
    return generator.generate_excel()
