"""
customer_extractor.py
=====================
Module to extract unique customers from Google Sheets with multiple year-based tabs.
Aggregates multiple phones/emails/names, saves registry, and exports Excel reports.
Reads all tabs where tab name matches year pattern (e.g., "2023", "2024", "2025").

Compatible with Streamlit apps.
"""

import re
import io
import gc
from datetime import datetime
from collections import Counter
from typing import Dict, List, Optional, Tuple, Set
import warnings

import pandas as pd
import numpy as np
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore', category=UserWarning)

# ==========================
#  MEMORY MANAGEMENT
# ==========================
class MemoryErrorHandler:
    """
    Handles memory allocation errors gracefully.
    Provides chunked processing, fallback modes, and ensures app continues running.
    """
    
    @staticmethod
    def check_memory_available(min_mb: int = 100) -> Tuple[bool, int]:
        """
        Check if minimum memory is available.
        
        Args:
            min_mb: Minimum required memory in MB
            
        Returns:
            Tuple of (has_enough_memory, available_mb)
        """
        try:
            import psutil
            memory = psutil.virtual_memory()
            available_mb = memory.available // (1024 * 1024)
            return available_mb >= min_mb, available_mb
        except ImportError:
            # psutil not available, assume memory is sufficient
            return True, 0
        except Exception:
            return True, 0
    
    @staticmethod
    def estimate_df_memory(df: pd.DataFrame) -> int:
        """
        Estimate DataFrame memory usage in MB.
        
        Args:
            df: DataFrame to estimate
            
        Returns:
            Estimated memory in MB
        """
        try:
            return df.memory_usage(deep=True).sum() // (1024 * 1024)
        except Exception:
            # Fallback estimation: ~100 bytes per cell
            return (len(df) * len(df.columns) * 100) // (1024 * 1024)
    
    @staticmethod
    def warn_if_low_memory(df: pd.DataFrame, operation: str = "operation"):
        """
        Warn user if memory appears low for the operation.
        """
        has_memory, available_mb = MemoryErrorHandler.check_memory_available(0)
        needed_mb = MemoryErrorHandler.estimate_df_memory(df) * 3  # 3x for processing overhead
        
        if has_memory and available_mb > 0 and available_mb < needed_mb:
            st.warning(
                f"⚠️ Low memory warning for {operation}: "
                f"~{available_mb}MB available, ~{needed_mb}MB recommended. "
                f"Processing may be slower or fail. Consider closing other applications."
            )
    
    @staticmethod
    def safe_concat(dfs: List[pd.DataFrame], chunk_size: int = 50000, 
                     on_error: str = "partial") -> Tuple[pd.DataFrame, Dict]:
        """
        Safely concatenate dataframes with memory protection.
        
        Args:
            dfs: List of dataframes to concatenate
            chunk_size: Maximum rows to process at once
            on_error: "partial" (return what we have), "empty" (return empty df), "raise" (re-raise)
        
        Returns:
            Tuple of (concatenated_df, metadata_dict)
        """
        if not dfs:
            return pd.DataFrame(), {"status": "empty", "rows": 0, "chunks": 0}
        
        metadata = {
            "total_input_dfs": len(dfs),
            "total_input_rows": sum(len(df) for df in dfs),
            "chunks": 0,
            "errors": [],
            "status": "success"
        }
        
        try:
            # Try normal concat first for small datasets
            total_rows = metadata["total_input_rows"]
            if total_rows < chunk_size:
                result = pd.concat(dfs, ignore_index=True)
                metadata["rows"] = len(result)
                return result, metadata
        except MemoryError as e:
            metadata["errors"].append(f"Initial concat failed: {str(e)}")
            st.warning("⚠️ Memory limit reached. Switching to chunked processing...")
        except Exception as e:
            metadata["errors"].append(f"Unexpected error in concat: {str(e)}")
        
        # Chunked concatenation for large datasets
        result_chunks = []
        current_chunk = []
        current_chunk_rows = 0
        
        for i, df in enumerate(dfs):
            try:
                df_rows = len(df)
                
                # If this single df exceeds chunk size, we need to process it in parts
                if df_rows > chunk_size:
                    st.info(f"📦 Processing large dataframe {i+1} in parts ({df_rows:,} rows)...")
                    for start_idx in range(0, df_rows, chunk_size):
                        try:
                            end_idx = min(start_idx + chunk_size, df_rows)
                            part = df.iloc[start_idx:end_idx].copy()
                            result_chunks.append(part)
                            metadata["chunks"] += 1
                            
                            # Force garbage collection between chunks
                            gc.collect()
                        except MemoryError as e:
                            metadata["errors"].append(f"Chunk {start_idx}-{end_idx} failed: {str(e)}")
                            st.warning(f"⚠️ Skipping rows {start_idx:,}-{end_idx:,} due to memory limit")
                            if on_error == "raise":
                                raise
                else:
                    # Check if adding this df would exceed chunk size
                    if current_chunk_rows + df_rows > chunk_size and current_chunk:
                        # Process current chunk
                        try:
                            combined_chunk = pd.concat(current_chunk, ignore_index=True)
                            result_chunks.append(combined_chunk)
                            metadata["chunks"] += 1
                            
                            # Reset chunk
                            current_chunk = [df]
                            current_chunk_rows = df_rows
                            
                            gc.collect()
                        except MemoryError as e:
                            metadata["errors"].append(f"Chunk concat failed: {str(e)}")
                            if on_error == "raise":
                                raise
                            # Try to salvage individual dfs
                            for single_df in current_chunk:
                                result_chunks.append(single_df)
                                metadata["chunks"] += 1
                            current_chunk = [df]
                            current_chunk_rows = df_rows
                    else:
                        current_chunk.append(df)
                        current_chunk_rows += df_rows
            except Exception as e:
                metadata["errors"].append(f"DataFrame {i} failed: {str(e)}")
                st.warning(f"⚠️ Could not process dataframe {i+1}, skipping...")
                if on_error == "raise":
                    raise
        
        # Process remaining chunk
        if current_chunk:
            try:
                combined_chunk = pd.concat(current_chunk, ignore_index=True)
                result_chunks.append(combined_chunk)
                metadata["chunks"] += 1
            except MemoryError as e:
                metadata["errors"].append(f"Final chunk concat failed: {str(e)}")
                # Add individually
                for single_df in current_chunk:
                    result_chunks.append(single_df)
                    metadata["chunks"] += 1
        
        # Final combination
        if not result_chunks:
            metadata["status"] = "empty"
            return pd.DataFrame(), metadata
        
        try:
            if len(result_chunks) == 1:
                final_result = result_chunks[0]
            else:
                final_result = pd.concat(result_chunks, ignore_index=True)
            metadata["rows"] = len(final_result)
            metadata["status"] = "partial" if metadata["errors"] else "success"
            return final_result, metadata
        except MemoryError as e:
            metadata["errors"].append(f"Final concat failed: {str(e)}")
            metadata["status"] = "failed"
            
            if on_error == "partial" and result_chunks:
                # Return largest chunk we have
                largest = max(result_chunks, key=len)
                st.warning(f"⚠️ Returning largest available chunk ({len(largest):,} rows)")
                metadata["rows"] = len(largest)
                return largest, metadata
            elif on_error == "empty":
                return pd.DataFrame(), metadata
            else:
                raise
    
    @staticmethod
    def safe_groupby(df: pd.DataFrame, group_col: str, agg_rules: Dict,
                     on_error: str = "partial") -> Tuple[pd.DataFrame, Dict]:
        """
        Safely perform groupby aggregation with memory protection.
        """
        metadata = {"status": "success", "errors": []}
        
        if df.empty:
            return df, metadata
        
        try:
            result = df.groupby(group_col).agg(agg_rules)
            metadata["rows"] = len(result)
            return result, metadata
        except MemoryError as e:
            metadata["errors"].append(f"Groupby failed: {str(e)}")
            st.warning("⚠️ Memory limit during grouping. Trying alternative approach...")
            
            if on_error == "partial":
                # Try processing in chunks by splitting the dataframe
                try:
                    unique_keys = df[group_col].unique()
                    chunk_size = max(1, len(unique_keys) // 4)  # Process 1/4 at a time
                    
                    result_parts = []
                    for i in range(0, len(unique_keys), chunk_size):
                        key_subset = unique_keys[i:i + chunk_size]
                        subset_df = df[df[group_col].isin(key_subset)]
                        try:
                            part_result = subset_df.groupby(group_col).agg(agg_rules)
                            result_parts.append(part_result)
                            gc.collect()
                        except MemoryError:
                            st.warning(f"⚠️ Skipping key subset {i}-{i+chunk_size}")
                            continue
                    
                    if result_parts:
                        final_result = pd.concat(result_parts)
                        metadata["rows"] = len(final_result)
                        metadata["status"] = "partial"
                        return final_result, metadata
                except Exception as e2:
                    metadata["errors"].append(f"Chunked groupby failed: {str(e2)}")
            
            if on_error == "empty":
                return pd.DataFrame(), metadata
            raise
    
    @staticmethod
    def safe_merge_registries(new_customers: pd.DataFrame, 
                              old_registry: Optional[pd.DataFrame],
                              chunk_size: int = 1000) -> Tuple[pd.DataFrame, Dict]:
        """
        Memory-safe registry merging with chunked processing.
        """
        metadata = {"status": "success", "new_count": len(new_customers), 
                   "old_count": len(old_registry) if old_registry is not None else 0,
                   "errors": []}
        
        if old_registry is None or old_registry.empty:
            new_customers = new_customers.copy()
            new_customers['first_seen'] = datetime.now()
            new_customers['last_updated'] = datetime.now()
            metadata["final_count"] = len(new_customers)
            return new_customers, metadata
        
        try:
            # Try normal merge first for small datasets
            if len(new_customers) < chunk_size and len(old_registry) < chunk_size * 2:
                result = merge_registries(new_customers, old_registry)
                metadata["final_count"] = len(result)
                return result, metadata
        except MemoryError:
            st.warning("⚠️ Memory limit during merge. Using chunked processing...")
        except Exception as e:
            metadata["errors"].append(f"Initial merge failed: {str(e)}")
        
        # Chunked processing
        # Prepare old registry lookup keys
        old_registry = old_registry.copy()
        old_registry['primary_email'] = old_registry['primary_email'].fillna('').astype(str)
        old_registry['primary_phone'] = old_registry['primary_phone'].fillna('').astype(str)
        old_registry['match_email'] = old_registry['primary_email'].str.lower()
        old_registry['match_phone'] = old_registry['primary_phone'].apply(normalize_phone)
        
        new_customers = new_customers.copy()
        new_customers['primary_email'] = new_customers['primary_email'].fillna('').astype(str)
        new_customers['primary_phone'] = new_customers['primary_phone'].fillna('').astype(str)
        new_customers['match_email'] = new_customers['primary_email'].str.lower()
        new_customers['match_phone'] = new_customers['primary_phone'].apply(normalize_phone)
        
        updated_rows = []
        total_rows = len(new_customers)
        
        # Create progress indicators
        progress_bar = st.progress(0)
        status_text = st.empty()
        start_time = datetime.now()
        
        processed_emails = set()
        processed_phones = set()
        
        for idx in range(0, total_rows, chunk_size):
            chunk_end = min(idx + chunk_size, total_rows)
            chunk = new_customers.iloc[idx:chunk_end].copy()
            
            try:
                for _, new_row in chunk.iterrows():
                    match_email = new_row['match_email']
                    match_phone = new_row['match_phone']
                    
                    # Skip if already processed this contact
                    if match_email and match_email in processed_emails:
                        continue
                    if match_phone and match_phone in processed_phones:
                        continue
                    
                    # Find match in old registry
                    old_match = old_registry[
                        ((old_registry['match_email'] == match_email) & (match_email != "")) |
                        ((old_registry['match_phone'] == match_phone) & (match_phone != ""))
                    ]
                    
                    if not old_match.empty:
                        # Update existing customer (simplified)
                        old_row = old_match.iloc[0].copy()
                        old_row['last_updated'] = datetime.now()
                        updated_rows.append(old_row)
                        
                        if match_email:
                            processed_emails.add(match_email)
                        if match_phone:
                            processed_phones.add(match_phone)
                    else:
                        # New customer
                        new_row_copy = new_row.copy()
                        new_row_copy['first_seen'] = datetime.now()
                        new_row_copy['last_updated'] = datetime.now()
                        updated_rows.append(new_row_copy)
                        
                        if match_email:
                            processed_emails.add(match_email)
                        if match_phone:
                            processed_phones.add(match_phone)
                
                # Update progress
                progress = chunk_end / total_rows
                progress_bar.progress(min(progress, 0.99))
                elapsed = (datetime.now() - start_time).total_seconds()
                rate = chunk_end / elapsed if elapsed > 0 else 0
                status_text.text(f"🔄 Merging: {chunk_end:,} / {total_rows:,} ({int(progress*100)}%) | {rate:.0f} rows/sec")
                
                # Force garbage collection
                gc.collect()
            except MemoryError as e:
                metadata["errors"].append(f"Chunk {idx}-{chunk_end} failed: {str(e)}")
                st.warning(f"⚠️ Memory issue at rows {idx:,}-{chunk_end:,}, some customers may be skipped")
                gc.collect()

        # Add remaining old customers who weren't updated
        for idx, row in old_reg_proc.iterrows():
            if idx not in processed_old_indices:
                updated_rows.append(row)

        progress_bar.progress(1.0)
        total_time = (datetime.now() - start_time).total_seconds()
        status_text.text(f"✅ Merged {len(updated_rows):,} customers in {total_time:.1f}s")
        
        # Build final DataFrame
        if updated_rows:
            final_df = pd.DataFrame(updated_rows)
            # Drop temporary columns
            for col in ['match_email', 'match_phone']:
                if col in final_df.columns:
                    final_df.drop(columns=[col], inplace=True)
            metadata["final_count"] = len(final_df)
            metadata["status"] = "partial" if metadata["errors"] else "success"
            return final_df, metadata
        else:
            metadata["status"] = "failed"
            return old_registry, metadata


def with_memory_protection(func, *args, fallback_value=None, **kwargs):
    """
    Wrapper to execute a function with memory error protection.
    
    Args:
        func: Function to execute
        fallback_value: Value to return if memory error occurs
        *args, **kwargs: Arguments to pass to func
    
    Returns:
        Tuple of (result, success_boolean, error_message)
    """
    gc.collect()  # Clear memory before operation
    
    try:
        result = func(*args, **kwargs)
        return result, True, None
    except MemoryError as e:
        gc.collect()
        error_msg = f"Memory allocation failed: {str(e)}"
        st.error(f"⚠️ {error_msg}")
        st.info("💡 Try: 1) Closing other applications, 2) Processing smaller dataset, 3) Restarting the app")
        if fallback_value is not None:
            return fallback_value, False, error_msg
        raise
    except Exception as e:
        error_msg = f"Error: {str(e)}"
        st.error(error_msg)
        if fallback_value is not None:
            return fallback_value, False, error_msg
        raise


# ==========================
#  CONFIGURATION
# ==========================
YEAR_PATTERN = re.compile(r'^\d{4}$')  # Matches 4-digit years like 2023, 2024, 2025

# Column detection patterns
_PHONE_PATTERNS = ["phone", "mobile", "contact", "cell", "telephone"]
_EMAIL_PATTERNS = ["email", "e-mail", "mail"]
_NAME_PATTERNS = ["name", "customer", "full name", "buyer", "billing name", "shipping name"]
_ORDER_PATTERNS = ["order number", "order id", "order#", "id", "order"]
_AMOUNT_PATTERNS = ["total", "amount", "price", "cost", "grand total"]
_DATE_PATTERNS = ["date", "order date", "created", "timestamp", "ordered", "purchase date"]


# ==========================
#  SESSION KEYS
# ==========================
_SESSION_KEY = "ce_df"
_SESSION_COLS = "ce_cols"
_SESSION_REGISTRY = "ce_registry"


# ==========================
#  HELPER FUNCTIONS
# ==========================
def normalize_phone(phone: str) -> str:
    """Extract digits only for phone normalization."""
    if pd.isna(phone) or not isinstance(phone, str):
        return ""
    return re.sub(r'\D', '', phone)


def normalize_email(email: str) -> str:
    """Lowercase and strip email."""
    if pd.isna(email) or not isinstance(email, str):
        return ""
    return email.strip().lower()


def _best_column(df: pd.DataFrame, patterns: List[str]) -> Optional[str]:
    """Return the first column whose lower-case name contains any pattern."""
    cols_lower = {c.lower(): c for c in df.columns}
    for pat in patterns:
        for lower_col, orig_col in cols_lower.items():
            if pat in lower_col:
                return orig_col
    return None


def detect_columns(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    """Auto-detect semantic column roles from the dataframe."""
    return {
        "phone": _best_column(df, _PHONE_PATTERNS),
        "email": _best_column(df, _EMAIL_PATTERNS),
        "name": _best_column(df, _NAME_PATTERNS),
        "order_id": _best_column(df, _ORDER_PATTERNS),
        "amount": _best_column(df, _AMOUNT_PATTERNS),
        "date": _best_column(df, _DATE_PATTERNS),
    }


# ==========================
#  GOOGLE SHEETS MULTI-TAB LOADING
# ==========================
def get_sheet_tabs(sheet_id: str, api_key: Optional[str] = None) -> List[Dict]:
    """
    Get all tab names from a Google Sheet using Sheets API.
    Returns list of dicts with 'title' and 'sheetId'.
    
    If no API key provided, falls back to CSV export attempt for public sheets.
    """
    if not api_key:
        # Try to detect tabs by attempting common gid values
        # This is a fallback for public sheets without API key
        return []
    
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}"
    params = {"key": api_key}
    
    try:
        response = requests.get(url, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
        sheets = data.get("sheets", [])
        return [{"title": s["properties"]["title"], "sheetId": s["properties"]["sheetId"]} 
                for s in sheets]
    except Exception as e:
        st.error(f"Failed to fetch sheet tabs: {e}")
        return []


def download_from_gid(sheet_id: str, gid: str, format: str = "csv") -> pd.DataFrame:
    """Download a specific tab from Google Sheets as CSV or TSV using gid."""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format={format}&gid={gid}"
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    
    # Auto-detect delimiter based on content
    content = response.content
    first_line = content.split(b'\n')[0] if b'\n' in content else content[:100]
    
    # Check for tab character in first line
    if b'\t' in first_line:
        return pd.read_csv(io.BytesIO(content), dtype=str, on_bad_lines="skip", sep='\t')
    else:
        return pd.read_csv(io.BytesIO(content), dtype=str, on_bad_lines="skip")


def extract_sheet_id_from_url(url: str) -> Optional[str]:
    """Extract the Google Sheets ID from a URL."""
    # Match patterns like /d/SHEET_ID/ or d/SHEET_ID/
    patterns = [
        r'/d/([a-zA-Z0-9_-]+)',
        r'^([a-zA-Z0-9_-]{30,})',  # Direct sheet ID
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return None


def detect_year_from_tab_name(title: str) -> Optional[str]:
    """Extract year from tab name like '2025 order list', '2024', etc."""
    if not title:
        return None
    # Look for 4-digit year pattern
    year_match = re.search(r'\b(20\d{2})\b', str(title))
    if year_match:
        year = year_match.group(1)
        if 2021 <= int(year) <= datetime.now().year:
            return year
    return None


def detect_year_from_data(df: pd.DataFrame, date_col: Optional[str] = None) -> Optional[str]:
    """Try to detect which year the data belongs to by looking at date columns."""
    if df.empty:
        return None
    
    # Try to find a date column
    date_candidates = []
    for col in df.columns:
        col_lower = str(col).lower()
        if any(keyword in col_lower for keyword in ['date', 'created', 'time', 'order']):
            date_candidates.append(col)
    
    for col in date_candidates:
        try:
            # Try parsing as datetime
            parsed = pd.to_datetime(df[col], errors='coerce')
            valid_dates = parsed.dropna()
            if len(valid_dates) > 0:
                # Get the most common year
                years = valid_dates.dt.year.value_counts()
                if len(years) > 0:
                    most_common_year = years.index[0]
                    if 2021 <= most_common_year <= datetime.now().year:
                        return str(most_common_year)
        except Exception:
            continue
    
    return None


def load_year_tabs_from_sheet(url_or_id: str, api_key: Optional[str] = None) -> Tuple[pd.DataFrame, List[str]]:
    """
    Load all tabs from a Google Sheet where tab name matches year pattern (4 digits).
    Covers years from 2021 to current year.
    Returns combined DataFrame and list of tab names loaded.
    """
    sheet_id = extract_sheet_id_from_url(url_or_id)
    if not sheet_id:
        raise ValueError("Could not extract sheet ID from URL")
    
    current_year = datetime.now().year
    target_years = [str(y) for y in range(2021, current_year + 1)]
    
    all_data = []
    loaded_tabs = []
    
    # Try API first if key provided
    if api_key:
        available_tabs = get_sheet_tabs(sheet_id, api_key)
        if available_tabs:
            st.info(f"📋 Found {len(available_tabs)} tabs via API")
            for tab in available_tabs:
                # Check if tab name contains a year (e.g., "2025 order list", "2024")
                year_from_title = detect_year_from_tab_name(tab["title"])
                if year_from_title and year_from_title in target_years:
                    try:
                        df = download_from_gid(sheet_id, str(tab["sheetId"]))
                        if not df.empty:
                            df["_source_tab"] = year_from_title
                            all_data.append(df)
                            loaded_tabs.append(year_from_title)
                            st.success(f"✅ Loaded year tab: {tab['title']} → {year_from_title}")
                    except Exception as e:
                        st.warning(f"⚠️ Failed to load tab '{tab['title']}': {e}")
    
    # If no data loaded yet, try discovery mode (common for public sheets)
    if not all_data:
        st.info("🔍 Discovering year tabs by trying gids (2021-{})...".format(current_year))
        
        # Common gid patterns: 0, 1, 2, 3... or specific large numbers
        # Try gids 0-15 first (most common for sequential tabs)
        gids_to_try = list(range(0, 20))  # Try 0-19
        
        for gid in gids_to_try:
            try:
                df = download_from_gid(sheet_id, str(gid))
                if not df.empty:
                    # Try to detect year from data first
                    detected_year = detect_year_from_data(df)
                    
                    # If detection failed, try gid-based mapping
                    if not detected_year:
                        # Assume gid 0 = 2021, gid 1 = 2022, etc.
                        year_idx = len(loaded_tabs)
                        if year_idx < len(target_years):
                            detected_year = target_years[year_idx]
                    
                    if detected_year and detected_year not in loaded_tabs:
                        df["_source_tab"] = detected_year
                        df["_gid"] = gid
                        all_data.append(df)
                        loaded_tabs.append(detected_year)
                        st.success(f"✅ Found year tab: {detected_year} (gid={gid})")
            except Exception as e:
                # Tab doesn't exist or no access
                continue
    
    # Final fallback: try common gid values that might be year tabs
    # Some sheets use specific gids like 1234567890 for different tabs
    if not all_data:
        st.info("🔍 Trying alternative gid patterns...")
        # Try extracting gid from the original URL if present
        url_match = re.search(r'[?&]gid=(\d+)', url_or_id)
        if url_match:
            gid_from_url = url_match.group(1)
            try:
                df = download_from_gid(sheet_id, gid_from_url)
                if not df.empty:
                    # Use tab name from data or default to the gid
                    tab_name = detect_year_from_data(df) or f"Tab_{gid_from_url}"
                    df["_source_tab"] = tab_name
                    df["_gid"] = gid_from_url
                    all_data.append(df)
                    loaded_tabs.append(tab_name)
                    st.success(f"✅ Loaded from URL gid: {tab_name}")
            except Exception:
                pass
    
    # If still no data, try loading gid 0 (default/first tab) as a last resort
    if not all_data:
        try:
            st.info("🔍 Trying to load default tab (gid=0)...")
            df = download_from_gid(sheet_id, "0")
            if not df.empty:
                df["_source_tab"] = "Default"
                df["_gid"] = 0
                all_data.append(df)
                loaded_tabs.append("Default")
                st.success(f"✅ Loaded default tab with {len(df):,} rows")
        except Exception:
            pass
    
    if not all_data:
        raise ValueError(
            "No year tabs (2021-{}) could be loaded. \n"
            "Please ensure:\n"
            "1. The Google Sheet is published to web (File → Share → Publish to web)\n"
            "2. Tabs are named by year (2021, 2022, etc.) or contain date columns\n"
            "3. The URL is a public share link".format(current_year)
        )
    
    # Use memory-safe concatenation
    combined, concat_meta = MemoryErrorHandler.safe_concat(all_data, chunk_size=50000, on_error="partial")
    
    if concat_meta["status"] == "failed":
        st.error("❌ Could not combine data due to memory limitations")
        raise MemoryError("Failed to concatenate data: " + "; ".join(concat_meta.get("errors", [])))
    elif concat_meta["status"] == "partial":
        st.warning(f"⚠️ Partial data loaded due to memory constraints. Using {len(combined):,} rows.")
    else:
        st.success(f"✅ Combined {len(loaded_tabs)} year tabs with {len(combined):,} total rows")
    
    if concat_meta.get("errors"):
        st.caption(f"Processing notes: {len(concat_meta['errors'])} chunks had issues")
    
    return combined, loaded_tabs


def load_from_url_simple(url: str) -> pd.DataFrame:
    """Download CSV/TSV from a URL (single tab). Auto-detects format."""
    # Convert Google Sheets /edit URL to /pub export URL if needed
    if "docs.google.com/spreadsheets" in url:
        if "output=csv" not in url and "output=tsv" not in url:
            # Check if user specified tsv in the URL they provided
            if "tsv" in url.lower():
                url = re.sub(r"/edit.*", "", url) + "/pub?output=tsv"
            else:
                url = re.sub(r"/edit.*", "", url) + "/pub?output=csv"
    
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    content = resp.content
    
    # Auto-detect delimiter
    first_line = content.split(b'\n')[0] if b'\n' in content else content[:100]
    if b'\t' in first_line or b'tsv' in url.lower().encode():
        return pd.read_csv(io.BytesIO(content), dtype=str, on_bad_lines="skip", sep='\t')
    else:
        return pd.read_csv(io.BytesIO(content), dtype=str, on_bad_lines="skip")


# ==========================
#  DATA CLEANING
# ==========================
def clean_dataframe(df: pd.DataFrame, cols: Dict[str, Optional[str]]) -> pd.DataFrame:
    """Rename detected columns to standard names and clean."""
    df = df.copy()
    
    # Normalize contact fields
    if cols.get("phone"):
        df["_phone_norm"] = df[cols["phone"]].apply(normalize_phone)
    else:
        df["_phone_norm"] = ""
    
    if cols.get("email"):
        df["_email_norm"] = df[cols["email"]].apply(normalize_email)
    else:
        df["_email_norm"] = ""
    
    if cols.get("name"):
        df["_name_clean"] = df[cols["name"]].fillna("").astype(str).str.strip()
    else:
        df["_name_clean"] = ""
    
    # Convert amount to numeric
    if cols.get("amount"):
        df["_amount"] = pd.to_numeric(df[cols["amount"]], errors="coerce").fillna(0)
    else:
        df["_amount"] = 0
    
    # Convert date
    if cols.get("date"):
        df["_date"] = pd.to_datetime(df[cols["date"]], errors="coerce")
    else:
        df["_date"] = pd.NaT
    
    # Create grouping key: email if exists, else phone, else None
    # Use .replace("", None) to convert empty strings to None/NaN
    df["_email_norm"] = df["_email_norm"].replace("", np.nan)
    df["_phone_norm"] = df["_phone_norm"].replace("", np.nan)
    
    # Debug: show counts
    email_count = df["_email_norm"].notna().sum()
    phone_count = df["_phone_norm"].notna().sum()
    st.write(f"📊 After cleaning: {email_count:,} with email, {phone_count:,} with phone")
    
    df["_group_key"] = df["_email_norm"]
    # Fill missing emails with phone
    df["_group_key"] = df["_group_key"].fillna(df["_phone_norm"])
    
    # Drop rows with no contact info (no email and no phone)
    df = df[df["_group_key"].notna()].copy()
    
    st.write(f"📊 Rows with contact info: {len(df):,}")
    
    # If dataframe is empty after filtering, add a dummy row to prevent downstream errors
    # but warn the user
    if df.empty:
        st.warning("⚠️ No rows with valid phone or email found. Check your column mappings.")
        # Return empty dataframe with expected columns
        df = pd.DataFrame(columns=["_phone_norm", "_email_norm", "_name_clean", "_amount", "_date", "_group_key"])
    
    return df


# ==========================
#  CUSTOMER GROUPING
# ==========================
def group_customers(df: pd.DataFrame, cols: Dict[str, Optional[str]]) -> pd.DataFrame:
    """
    Group rows by unique customer (using email_norm or phone_norm).
    Aggregates: all phones, emails, names, order_ids, total spent, first/last date.
    Also determines primary/secondary phone, email, name.
    """
    # Handle empty dataframe
    if df.empty:
        st.warning("⚠️ No customer data to group. Returning empty registry.")
        return pd.DataFrame(columns=[
            "customer_id", "primary_phone", "secondary_phones", "primary_email", 
            "secondary_emails", "primary_name", "all_names", "order_ids", 
            "order_count", "total_spent", "first_order_date", "last_order_date", "source_years"
        ])
    
    def collect_distinct(series: pd.Series) -> List[str]:
        return sorted(list(set([str(x) for x in series if pd.notna(x) and str(x) != ""])))
    
    def collect_names(series: pd.Series) -> List[str]:
        return [str(x) for x in series if pd.notna(x) and str(x) != ""]
    
    def mode_name(names_list: List[str]) -> str:
        if not names_list:
            return ""
        counter = Counter(names_list)
        return counter.most_common(1)[0][0]
    
    def combine_order_ids(series: pd.Series) -> str:
        ids = [str(x) for x in series if pd.notna(x) and str(x) != ""]
        return ", ".join(sorted(set(ids)))
    
    # Always use normalized columns for aggregation
    # These are guaranteed to exist after clean_dataframe
    phone_col = "_phone_norm"
    email_col = "_email_norm"
    name_col = "_name_clean"
    order_col = cols.get("order_id")
    
    # Build aggregation rules only for columns that exist
    agg_rules = {
        "_amount": "sum",
        "_date": ["min", "max"],
    }
    
    # Only add aggregation for contact columns if they exist and have data
    if phone_col in df.columns and df[phone_col].notna().any():
        agg_rules[phone_col] = collect_distinct
    if email_col in df.columns and df[email_col].notna().any():
        agg_rules[email_col] = collect_distinct
    if name_col in df.columns and df[name_col].notna().any():
        agg_rules[name_col] = collect_names
    
    if order_col and order_col in df.columns:
        agg_rules[order_col] = combine_order_ids
    
    # Use memory-safe groupby
    grouped, groupby_meta = MemoryErrorHandler.safe_groupby(
        df, "_group_key", agg_rules, on_error="partial"
    )
    
    if groupby_meta["status"] == "failed":
        st.error("❌ Could not group customers due to memory limitations")
        raise MemoryError("Failed to group customers: " + "; ".join(groupby_meta.get("errors", [])))
    elif groupby_meta["status"] == "partial":
        st.warning(f"⚠️ Partial customer grouping due to memory constraints. Processed {len(grouped):,} groups.")
    
    # Debug: show columns
    st.write("📊 Grouped columns:", list(grouped.columns))
    
    # Flatten multi-level column names
    if isinstance(grouped.columns, pd.MultiIndex):
        grouped.columns = ['_'.join(col).strip().rstrip('_') for col in grouped.columns.values]
        st.write("📊 Flattened columns:", list(grouped.columns))
    
    grouped = grouped.reset_index()
    
    # Build final columns
    final_df = pd.DataFrame()
    
    # Get _group_key from grouped (it was reset_index so it's now a column)
    if "_group_key" in grouped.columns:
        final_df["_group_key"] = grouped["_group_key"]
    
    # Primary phone (first in sorted list, typically the most complete)
    # Handle both single columns and lists from aggregation
    if phone_col in grouped.columns:
        phone_series = grouped[phone_col]
        # Check if values are already lists or need to be converted
        final_df["primary_phone"] = phone_series.apply(
            lambda x: x[0] if isinstance(x, list) and len(x) > 0 else (str(x) if pd.notna(x) else "")
        )
        final_df["secondary_phones"] = phone_series.apply(
            lambda x: ", ".join(x[1:]) if isinstance(x, list) and len(x) > 1 else ""
        )
    else:
        final_df["primary_phone"] = ""
        final_df["secondary_phones"] = ""
    
    # Primary email (first in sorted list)
    if email_col in grouped.columns:
        email_series = grouped[email_col]
        final_df["primary_email"] = email_series.apply(
            lambda x: x[0] if isinstance(x, list) and len(x) > 0 else (str(x) if pd.notna(x) else "")
        )
        final_df["secondary_emails"] = email_series.apply(
            lambda x: ", ".join(x[1:]) if isinstance(x, list) and len(x) > 1 else ""
        )
    else:
        final_df["primary_email"] = ""
        final_df["secondary_emails"] = ""
    
    # Name handling - use mode for primary name
    if name_col in grouped.columns:
        name_series = grouped[name_col]
        final_df["primary_name"] = name_series.apply(
            lambda x: mode_name(x) if isinstance(x, list) and len(x) > 0 else (str(x) if pd.notna(x) else "")
        )
        final_df["all_names"] = name_series.apply(
            lambda x: ", ".join(set(x)) if isinstance(x, list) and len(x) > 0 else (str(x) if pd.notna(x) else "")
        )
    else:
        final_df["primary_name"] = ""
        final_df["all_names"] = ""
    
    # Order info
    if order_col and order_col in df.columns:
        order_ids_col = grouped.get(order_col, ["" for _ in range(len(grouped))])
        final_df["order_ids"] = order_ids_col
        final_df["order_count"] = [len(str(ids).split(", ")) if ids else 0 
                                    for ids in order_ids_col]
    else:
        final_df["order_ids"] = ""
        final_df["order_count"] = 1
    
    # Amount and dates
    amount_col = "_amount_sum" if "_amount_sum" in grouped.columns else "_amount"
    final_df["total_spent"] = grouped.get(amount_col, 0).round(2)
    
    date_min_col = "_date_min" if "_date_min" in grouped.columns else "_date"
    date_max_col = "_date_max" if "_date_max" in grouped.columns else "_date"
    final_df["first_order_date"] = grouped.get(date_min_col, pd.NaT)
    final_df["last_order_date"] = grouped.get(date_max_col, pd.NaT)
    
    # Add customer ID
    final_df["customer_id"] = range(1, len(final_df) + 1)
    
    # Add source tabs (which years this customer appears in)
    if "_source_tab" in df.columns and "_group_key" in final_df.columns:
        source_tabs = df.groupby("_group_key")["_source_tab"].apply(
            lambda x: ", ".join(sorted(set(x)))
        ).reset_index()
        source_tabs.columns = ["_group_key", "source_years"]
        final_df = final_df.merge(source_tabs, on="_group_key", how="left")
    
    # Drop _group_key from final output (internal use only)
    if "_group_key" in final_df.columns:
        final_df = final_df.drop(columns=["_group_key"])
    
    return final_df


# ==========================
#  REGISTRY PERSISTENCE
# ==========================
def save_registry(df: pd.DataFrame, filepath: str = "customer_registry.xlsx"):
    """Save unique customers registry to Excel."""
    try:
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='UniqueCustomers', index=False)
            
            # Add metadata sheet
            metadata = pd.DataFrame({
                'info': ['last_updated', 'total_unique_customers', 'export_tool'],
                'value': [datetime.now().isoformat(), len(df), 'CustomerExtractor v1']
            })
            metadata.to_excel(writer, sheet_name='Metadata', index=False)
        return True
    except Exception as e:
        st.error(f"Error saving registry: {e}")
        return False


def load_registry(filepath: str = "customer_registry.xlsx") -> Optional[pd.DataFrame]:
    """Load previously saved registry."""
    try:
        df = pd.read_excel(filepath, sheet_name='UniqueCustomers', engine='openpyxl')
        return df
    except FileNotFoundError:
        return None
    except Exception as e:
        st.error(f"Error loading registry: {e}")
        return None


def merge_registries(new_customers: pd.DataFrame, old_registry: Optional[pd.DataFrame]) -> pd.DataFrame:
    """
    Merge new unique customers with existing registry.
    Matching based on primary_email (lower) or normalized primary_phone.
    Updates existing entries with new orders/info, adds new ones.
    """
    if old_registry is None or old_registry.empty:
        new_customers['first_seen'] = datetime.now()
        new_customers['last_updated'] = datetime.now()
        return new_customers
    
    # Create normalized match keys in old registry
    old_registry = old_registry.copy()
    # Ensure string type before using .str accessor
    old_registry['primary_email'] = old_registry['primary_email'].fillna('').astype(str)
    old_registry['primary_phone'] = old_registry['primary_phone'].fillna('').astype(str)
    old_registry['match_email'] = old_registry['primary_email'].str.lower()
    old_registry['match_phone'] = old_registry['primary_phone'].apply(normalize_phone)
    
    # Prepare new customers
    new_customers = new_customers.copy()
    # Ensure string type before using .str accessor
    new_customers['primary_email'] = new_customers['primary_email'].fillna('').astype(str)
    new_customers['primary_phone'] = new_customers['primary_phone'].fillna('').astype(str)
    new_customers['match_email'] = new_customers['primary_email'].str.lower()
    new_customers['match_phone'] = new_customers['primary_phone'].apply(normalize_phone)
    
    updated_rows = []
    total_rows = len(new_customers)
    
    # Create progress bar
    progress_bar = st.progress(0)
    status_text = st.empty()
    start_time = datetime.now()
    
    for idx, new_row in new_customers.iterrows():
        # Update progress every 50 rows or at start/end
        if idx % 50 == 0 or idx == total_rows - 1:
            progress = (idx + 1) / total_rows
            progress_bar.progress(min(progress, 0.99))
            
            # Calculate ETA
            elapsed = (datetime.now() - start_time).total_seconds()
            rate = 0
            eta_str = "calculating..."
            if idx > 0 and elapsed > 0:
                rate = idx / elapsed  # rows per second
                remaining_rows = total_rows - idx
                eta_seconds = remaining_rows / rate if rate > 0 else 0
                eta_str = f"~{int(eta_seconds)}s" if eta_seconds < 60 else f"~{int(eta_seconds/60)}m"
            
            status_text.text(f"🔄 Merging: {idx+1:,} / {total_rows:,} customers ({int(progress*100)}%) | {rate:.0f} rows/sec | ETA: {eta_str}")
        
        match_email = new_row['match_email']
        match_phone = new_row['match_phone']
        
        # Find match in old registry
        old_match = old_registry[
            ((old_registry['match_email'] == match_email) & (match_email != "")) |
            ((old_registry['match_phone'] == match_phone) & (match_phone != ""))
        ]
        
        if not old_match.empty:
            # Update existing customer
            old_row = old_match.iloc[0].copy()
            
            # Merge phone lists
            existing_phones = set()
            if pd.notna(old_row.get('secondary_phones')):
                existing_phones.update(str(old_row['secondary_phones']).split(", "))
            if pd.notna(old_row.get('primary_phone')) and old_row['primary_phone']:
                existing_phones.add(str(old_row['primary_phone']))
            
            new_phones = set()
            if pd.notna(new_row.get('secondary_phones')):
                new_phones.update(str(new_row['secondary_phones']).split(", "))
            if pd.notna(new_row.get('primary_phone')) and new_row['primary_phone']:
                new_phones.add(str(new_row['primary_phone']))
            
            all_phones = sorted(list(existing_phones | new_phones))
            old_row['primary_phone'] = all_phones[0] if all_phones else ""
            old_row['secondary_phones'] = ", ".join(all_phones[1:]) if len(all_phones) > 1 else ""
            
            # Merge emails
            existing_emails = set()
            if pd.notna(old_row.get('secondary_emails')):
                existing_emails.update(str(old_row['secondary_emails']).split(", "))
            if pd.notna(old_row.get('primary_email')) and old_row['primary_email']:
                existing_emails.add(str(old_row['primary_email']))
            
            new_emails = set()
            if pd.notna(new_row.get('secondary_emails')):
                new_emails.update(str(new_row['secondary_emails']).split(", "))
            if pd.notna(new_row.get('primary_email')) and new_row['primary_email']:
                new_emails.add(str(new_row['primary_email']))
            
            all_emails = sorted(list(existing_emails | new_emails))
            old_row['primary_email'] = all_emails[0] if all_emails else ""
            old_row['secondary_emails'] = ", ".join(all_emails[1:]) if len(all_emails) > 1 else ""
            
            # Merge names
            existing_names = set()
            if pd.notna(old_row.get('all_names')):
                existing_names.update(str(old_row['all_names']).split(", "))
            new_names = set()
            if pd.notna(new_row.get('all_names')):
                new_names.update(str(new_row['all_names']).split(", "))
            all_names = sorted(list(existing_names | new_names))
            old_row['all_names'] = ", ".join(all_names)
            old_row['primary_name'] = all_names[0] if all_names else ""
            
            # Merge source years
            if 'source_years' in old_row and 'source_years' in new_row:
                old_years = set(str(old_row['source_years']).split(", ")) if pd.notna(old_row['source_years']) else set()
                new_years = set(str(new_row['source_years']).split(", ")) if pd.notna(new_row['source_years']) else set()
                old_row['source_years'] = ", ".join(sorted(old_years | new_years))
            
            # Merge order IDs if present
            if 'order_ids' in old_row and 'order_ids' in new_row:
                old_orders = set(str(old_row['order_ids']).split(", ")) if pd.notna(old_row['order_ids']) else set()
                new_orders = set(str(new_row['order_ids']).split(", ")) if pd.notna(new_row['order_ids']) else set()
                all_orders = sorted(old_orders | new_orders)
                old_row['order_ids'] = ", ".join(all_orders)
                old_row['order_count'] = len(all_orders)
            
            # Update dates and totals
            if pd.notna(old_row.get('total_spent')) and pd.notna(new_row.get('total_spent')):
                old_row['total_spent'] = float(old_row['total_spent']) + float(new_row['total_spent'])
            
            if pd.notna(old_row.get('first_order_date')) and pd.notna(new_row.get('first_order_date')):
                old_row['first_order_date'] = min(
                    pd.to_datetime(old_row['first_order_date']), 
                    pd.to_datetime(new_row['first_order_date'])
                )
            
            if pd.notna(old_row.get('last_order_date')) and pd.notna(new_row.get('last_order_date')):
                old_row['last_order_date'] = max(
                    pd.to_datetime(old_row['last_order_date']), 
                    pd.to_datetime(new_row['last_order_date'])
                )
            
            old_row['last_updated'] = datetime.now()
            updated_rows.append(old_row)
        else:
            # New customer
            new_row = new_row.copy()
            new_row['first_seen'] = datetime.now()
            new_row['last_updated'] = datetime.now()
            updated_rows.append(new_row)
    
    # Complete progress
    progress_bar.progress(1.0)
    total_time = (datetime.now() - start_time).total_seconds()
    status_text.text(f"✅ Merged {total_rows:,} customers in {total_time:.1f}s ({total_rows/total_time:.0f} rows/sec)")
    
    # Build final DataFrame
    final_df = pd.DataFrame(updated_rows)
    
    # Drop temporary columns
    for col in ['match_email', 'match_phone']:
        if col in final_df.columns:
            final_df.drop(columns=[col], inplace=True)
    
    return final_df


# ==========================
#  EXPORT REPORT
# ==========================
def export_full_report(customers_df: pd.DataFrame, raw_df: Optional[pd.DataFrame] = None,
                       filename: str = "customer_report.xlsx"):
    """
    Export a rich Excel report with styling.
    Sheets: Unique Customers, Raw Data (optional), Summary.
    """
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Unique Customers sheet
            customers_df.to_excel(writer, sheet_name='Unique Customers', index=False)
            
            # Raw Data sheet (if provided)
            if raw_df is not None:
                raw_df.to_excel(writer, sheet_name='Raw Source Data', index=False)
            
            # Summary sheet
            total = len(customers_df)
            summary_data = {
                'Metric': ['Total Unique Customers', 'With Phone', 'With Email', 'With Both', 'Export Date'],
                'Value': [
                    total,
                    (customers_df['primary_phone'] != "").sum(),
                    (customers_df['primary_email'] != "").sum(),
                    ((customers_df['primary_phone'] != "") & (customers_df['primary_email'] != "")).sum(),
                    datetime.now().strftime("%Y-%m-%d %H:%M")
                ]
            }
            summary = pd.DataFrame(summary_data)
            summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Apply styling
            workbook = writer.book
            for sheetname in workbook.sheetnames:
                ws = workbook[sheetname]
                
                # Header style
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Auto-size columns
                for column_cells in ws.columns:
                    max_length = 0
                    column_letter = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        return True
    except Exception as e:
        st.error(f"Error exporting report: {e}")
        return False


# ==========================
#  MAIN PIPELINE
# ==========================
def extract_customers_from_google_sheet(url: str,
                                        save_registry_path: str = "customer_registry.xlsx") -> Tuple[pd.DataFrame, Dict, pd.DataFrame]:
    """
    Complete pipeline:
    1. Download CSV from URL
    2. Clean and group customers
    3. Merge with existing registry (if any)
    4. Save updated registry
    5. Return final customer DataFrame and metadata
    """
    # Load data from all year tabs (2021 to current year)
    st.info("📥 Loading data from year tabs (2021-{})...".format(datetime.now().year))
    try:
        raw_df, loaded_tabs = load_year_tabs_from_sheet(url)
        st.success(f"✅ Loaded {len(loaded_tabs)} year tabs: {', '.join(loaded_tabs)}")
        st.success(f"✅ Total {len(raw_df):,} rows from all years.")
        
        # Memory warning for large datasets
        if len(raw_df) > 50000:
            MemoryErrorHandler.warn_if_low_memory(raw_df, "customer processing")
            st.info(f"💾 Loaded data size: ~{MemoryErrorHandler.estimate_df_memory(raw_df)}MB. Using chunked processing if needed.")
    except Exception as e:
        st.warning(f"Year tab loading failed: {e}. Falling back to single tab...")
        try:
            raw_df = load_from_url_simple(url)
            loaded_tabs = ["Single Tab"]
            st.success(f"✅ Loaded {len(raw_df):,} rows from single tab.")
        except Exception as e2:
            st.error(f"Failed to load from URL: {e2}")
            raise
    
    # Detect columns
    st.info("🔍 Detecting columns...")
    cols = detect_columns(raw_df)
    st.write("Detected columns:", {k: v for k, v in cols.items() if v})
    
    # Clean data
    cleaned_df = clean_dataframe(raw_df, cols)
    st.success(f"✅ Cleaned {len(cleaned_df):,} rows with contact info.")
    
    # Group customers
    st.info("👥 Grouping unique customers...")
    unique_customers = group_customers(cleaned_df, cols)
    st.success(f"✅ Found {len(unique_customers):,} unique customers.")
    
    # Load existing registry
    old_registry = load_registry(save_registry_path)
    
    # Merge registries with memory protection
    st.info("🔄 Merging with existing registry...")
    merged_customers, merge_meta = MemoryErrorHandler.safe_merge_registries(
        unique_customers, old_registry, chunk_size=1000
    )
    
    if merge_meta["status"] == "failed":
        st.error("❌ Could not merge registries due to memory limitations")
        st.warning("⚠️ Using new customers only, without merging with existing registry")
        merged_customers = unique_customers.copy()
        merged_customers['first_seen'] = datetime.now()
        merged_customers['last_updated'] = datetime.now()
    
    # Count truly new customers (those in unique_customers not in old_registry)
    old_count = merge_meta.get("old_count", 0)
    final_count = merge_meta.get("final_count", len(merged_customers))
    new_count = max(0, final_count - old_count)
    st.success(f"✅ Total unique customers: {len(merged_customers):,} (new: {new_count:,})")
    
    # Save registry
    if save_registry(merged_customers, save_registry_path):
        st.success(f"💾 Registry saved to `{save_registry_path}`")
    
    metadata = {
        "total_unique_customers": len(merged_customers),
        "new_customers_added": max(0, new_count),
        "last_run": datetime.now().isoformat(),
        "source_urls": [url], # Fix: reference correct variable
        "loaded_tabs": loaded_tabs,
        "memory_status": merge_meta.get("status", "unknown")
    }
    
    return merged_customers, metadata, raw_df


def extract_customers_from_year_urls(year_urls: Dict[str, str],
                                     save_registry_path: str = "customer_registry.xlsx") -> Tuple[pd.DataFrame, Dict, pd.DataFrame]:
    """
    Load customers from individual year URLs (one per year).
    year_urls = {"2021": "url1", "2022": "url2", ...}
    """
    st.info(f"📥 Loading data from {len(year_urls)} year URLs...")
    
    all_data = []
    loaded_tabs = []
    
    for year, url in sorted(year_urls.items()):
        try:
            st.info(f"  Loading {year}...")
            df = load_from_url_simple(url)
            if not df.empty:
                df["_source_tab"] = year
                all_data.append(df)
                loaded_tabs.append(year)
                st.success(f"  ✅ {year}: {len(df):,} rows")
        except Exception as e:
            st.warning(f"  ⚠️ Failed to load {year}: {e}")
            continue
    
    if not all_data:
        raise ValueError("No year data could be loaded from any of the provided URLs")
    
    # Use memory-safe concatenation
    raw_df, concat_meta = MemoryErrorHandler.safe_concat(all_data, chunk_size=50000, on_error="partial")
    
    if concat_meta["status"] == "failed":
        st.error("❌ Could not combine year data due to memory limitations")
        raise MemoryError("Failed to concatenate year data: " + "; ".join(concat_meta.get("errors", [])))
    elif concat_meta["status"] == "partial":
        st.warning(f"⚠️ Partial year data loaded due to memory constraints. Using {len(raw_df):,} rows.")
    else:
        st.success(f"✅ Combined {len(loaded_tabs)} years with {len(raw_df):,} total rows")
    
    # Detect columns
    st.info("🔍 Detecting columns...")
    cols = detect_columns(raw_df)
    st.write("Detected columns:", {k: v for k, v in cols.items() if v})
    
    # Clean data
    cleaned_df = clean_dataframe(raw_df, cols)
    st.success(f"✅ Cleaned {len(cleaned_df):,} rows with contact info.")
    
    # Group customers
    st.info("👥 Grouping unique customers...")
    unique_customers = group_customers(cleaned_df, cols)
    st.success(f"✅ Found {len(unique_customers):,} unique customers.")
    
    # Load existing registry
    old_registry = load_registry(save_registry_path)
    
    # Merge registries
    st.info("🔄 Merging with existing registry...")
    merged_customers = merge_registries(unique_customers, old_registry)
    # Count truly new customers (those in unique_customers not in old_registry)
    old_count = len(old_registry) if old_registry is not None else 0
    new_count = max(0, len(merged_customers) - old_count)
    st.success(f"✅ Total unique customers: {len(merged_customers):,} (new: {new_count:,})")
    
    # Save registry
    if save_registry(merged_customers, save_registry_path):
        st.success(f"💾 Registry saved to `{save_registry_path}`")
    
    metadata = {
        "total_unique_customers": len(merged_customers),
        "new_customers_added": max(0, new_count),
        "last_run": datetime.now().isoformat(),
        "source_urls": year_urls,
        "loaded_tabs": loaded_tabs
    }
    
    return merged_customers, metadata, raw_df


def extract_customers_from_single_tab(url: str,
                                      save_registry_path: str = "customer_registry.xlsx") -> Tuple[pd.DataFrame, Dict, pd.DataFrame]:
    """
    Extract customers from a single tab (non-year based tabs like 'Todays Sales', 'Last Day Sales').
    Does NOT try to discover year tabs - just loads the specified URL directly.
    """
    st.info(f"📥 Loading single tab from URL...")
    
    try:
        raw_df = load_from_url_simple(url)
        tab_name = "SingleTab"
        
        # Try to detect tab name from URL gid if present
        url_match = re.search(r'[?&]gid=(\d+)', url)
        if url_match:
            tab_name = f"Tab_{url_match.group(1)}"
        
        if not raw_df.empty:
            raw_df["_source_tab"] = tab_name
            st.success(f"✅ Loaded tab '{tab_name}' with {len(raw_df):,} rows")
        else:
            raise ValueError("Loaded data is empty")
    except Exception as e:
        st.error(f"Failed to load from URL: {e}")
        raise
    
    # Detect columns
    st.info("🔍 Detecting columns...")
    cols = detect_columns(raw_df)
    st.write("Detected columns:", {k: v for k, v in cols.items() if v})
    
    # Clean data
    cleaned_df = clean_dataframe(raw_df, cols)
    st.success(f"✅ Cleaned {len(cleaned_df):,} rows with contact info.")
    
    # Group customers
    st.info("👥 Grouping unique customers...")
    unique_customers = group_customers(cleaned_df, cols)
    st.success(f"✅ Found {len(unique_customers):,} unique customers.")
    
    # Load existing registry
    old_registry = load_registry(save_registry_path)
    
    # Merge registries with memory protection
    st.info("🔄 Merging with existing registry...")
    merged_customers, merge_meta = MemoryErrorHandler.safe_merge_registries(
        unique_customers, old_registry, chunk_size=1000
    )
    
    if merge_meta["status"] == "failed":
        st.error("❌ Could not merge registries due to memory limitations")
        st.warning("⚠️ Using new customers only, without merging with existing registry")
        merged_customers = unique_customers.copy()
        merged_customers['first_seen'] = datetime.now()
        merged_customers['last_updated'] = datetime.now()
    
    # Count truly new customers (those in unique_customers not in old_registry)
    old_count = merge_meta.get("old_count", 0)
    final_count = merge_meta.get("final_count", len(merged_customers))
    new_count = max(0, final_count - old_count)
    st.success(f"✅ Total unique customers: {len(merged_customers):,} (new: {new_count:,})")
    
    # Save registry
    if save_registry(merged_customers, save_registry_path):
        st.success(f"💾 Registry saved to `{save_registry_path}`")
    
    metadata = {
        "total_unique_customers": len(merged_customers),
        "new_customers_added": max(0, new_count),
        "last_run": datetime.now().isoformat(),
        "source_url": url,
        "loaded_tabs": [tab_name],
        "memory_status": merge_meta.get("status", "unknown")
    }
    
    return merged_customers, metadata, raw_df


def extract_customers_from_uploaded_files(uploaded_files: Dict[str, any],
                                          save_registry_path: str = "customer_registry.xlsx") -> Tuple[pd.DataFrame, Dict, pd.DataFrame]:
    """
    Load customers from uploaded files (CSV, TSV, Excel).
    uploaded_files = {"2021": FileObject, "2022": FileObject, ...}
    Auto-detects file format from content.
    """
    st.info(f"📥 Loading data from {len(uploaded_files)} files...")
    
    all_data = []
    loaded_tabs = []
    
    for year, file_obj in sorted(uploaded_files.items()):
        try:
            st.info(f"  Loading {year} ({file_obj.name})...")
            
            # Read file content
            content = file_obj.read()
            file_obj.seek(0)  # Reset for potential re-read
            
            # Auto-detect format
            file_name = file_obj.name.lower()
            
            if file_name.endswith(('.xlsx', '.xls')):
                # Excel file
                df = pd.read_excel(io.BytesIO(content), dtype=str)
            elif file_name.endswith('.tsv') or file_name.endswith('.txt'):
                # TSV file
                df = pd.read_csv(io.BytesIO(content), dtype=str, on_bad_lines="skip", sep='\t')
            else:
                # CSV or auto-detect
                first_line = content.split(b'\n')[0] if b'\n' in content else content[:100]
                if b'\t' in first_line:
                    df = pd.read_csv(io.BytesIO(content), dtype=str, on_bad_lines="skip", sep='\t')
                else:
                    df = pd.read_csv(io.BytesIO(content), dtype=str, on_bad_lines="skip")
            
            if not df.empty:
                df["_source_tab"] = year
                all_data.append(df)
                loaded_tabs.append(year)
                st.success(f"  ✅ {year}: {len(df):,} rows")
        except Exception as e:
            st.warning(f"  ⚠️ Failed to load {year}: {e}")
            continue
    
    if not all_data:
        raise ValueError("No data could be loaded from any of the uploaded files")
    
    # Use memory-safe concatenation
    raw_df, concat_meta = MemoryErrorHandler.safe_concat(all_data, chunk_size=50000, on_error="partial")
    
    if concat_meta["status"] == "failed":
        st.error("❌ Could not combine file data due to memory limitations")
        raise MemoryError("Failed to concatenate file data: " + "; ".join(concat_meta.get("errors", [])))
    elif concat_meta["status"] == "partial":
        st.warning(f"⚠️ Partial file data loaded due to memory constraints. Using {len(raw_df):,} rows.")
    else:
        st.success(f"✅ Combined {len(loaded_tabs)} files with {len(raw_df):,} total rows")
    
    # Detect columns
    st.info("🔍 Detecting columns...")
    cols = detect_columns(raw_df)
    st.write("Detected columns:", {k: v for k, v in cols.items() if v})
    
    # Clean data
    cleaned_df = clean_dataframe(raw_df, cols)
    st.success(f"✅ Cleaned {len(cleaned_df):,} rows with contact info.")
    
    # Group customers
    st.info("👥 Grouping unique customers...")
    unique_customers = group_customers(cleaned_df, cols)
    st.success(f"✅ Found {len(unique_customers):,} unique customers.")
    
    # Load existing registry
    old_registry = load_registry(save_registry_path)
    
    # Merge registries with memory protection
    st.info("🔄 Merging with existing registry...")
    merged_customers, merge_meta = MemoryErrorHandler.safe_merge_registries(
        unique_customers, old_registry, chunk_size=1000
    )
    
    if merge_meta["status"] == "failed":
        st.error("❌ Could not merge registries due to memory limitations")
        st.warning("⚠️ Using new customers only, without merging with existing registry")
        merged_customers = unique_customers.copy()
        merged_customers['first_seen'] = datetime.now()
        merged_customers['last_updated'] = datetime.now()
    
    # Count truly new customers (those in unique_customers not in old_registry)
    old_count = merge_meta.get("old_count", 0)
    final_count = merge_meta.get("final_count", len(merged_customers))
    new_count = max(0, final_count - old_count)
    st.success(f"✅ Total unique customers: {len(merged_customers):,} (new: {new_count:,})")
    
    # Save registry
    if save_registry(merged_customers, save_registry_path):
        st.success(f"💾 Registry saved to `{save_registry_path}`")
    
    metadata = {
        "total_unique_customers": len(merged_customers),
        "new_customers_added": max(0, new_count),
        "last_run": datetime.now().isoformat(),
        "source_files": {y: f.name for y, f in uploaded_files.items()},
        "loaded_tabs": loaded_tabs,
        "memory_status": merge_meta.get("status", "unknown")
    }
    
    return merged_customers, metadata, raw_df


# ==========================
#  STREAMLIT UI
# ==========================
def _metric_card(col, label: str, value: str, icon: str = ""):
    """Render a styled metric card."""
    col.markdown(
        f"""
        <div style="
            background: linear-gradient(135deg,#1e293b,#0f172a);
            border:1px solid #334155;
            border-radius:12px;
            padding:18px 20px;
            text-align:center;
        ">
            <div style="font-size:1.8rem;">{icon}</div>
            <div style="font-size:1.9rem;font-weight:700;color:#38bdf8;">{value}</div>
            <div style="font-size:0.82rem;color:#94a3b8;margin-top:4px;">{label}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_customer_extractor_tab():
    """Main Streamlit render entry-point for the Customer Extractor tab."""
    
    st.markdown(
        """
        <style>
        .ce-header{
            background:linear-gradient(90deg,#0ea5e9,#6366f1);
            -webkit-background-clip:text;-webkit-text-fill-color:transparent;
            font-size:1.7rem;font-weight:800;margin-bottom:.2rem;
        }
        .ce-sub{color:#94a3b8;font-size:.9rem;margin-bottom:1.2rem;}
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown('<div class="ce-header">📊 Customer Data Extractor</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="ce-sub">Extract unique customers from Google Sheets. '
        'Aggregates multiple phones/emails, maintains persistent registry.</div>',
        unsafe_allow_html=True,
    )
    
    # ── Data Source ───────────────────────────────────────────────────────
    st.markdown("#### 🌐 Google Sheets URL")
    
    default_url = (
        "https://docs.google.com/spreadsheets/d/e/"
        "2PACX-1vTBDukmkRJGgHjCRIAAwGmlWaiPwESXSp9UBXm3_sbs37bk2HxavPc62aobmL1cGWUfAKE4Zd6yJySO"
        "/pub?output=csv"
    )
    
    # Input mode: Single URL vs Multiple Year URLs vs File Upload
    input_mode = st.radio(
        "Input Mode",
        ["Single URL (Auto-discover year tabs)", "Single URL (One specific tab)", "Multiple URLs (One per year)", "Upload File(s)"],
        horizontal=True,
        key="ce_input_mode"
    )
    
    year_urls = {}
    uploaded_files = {}
    url_input = None
    single_tab_mode = False
    
    if input_mode == "Single URL (Auto-discover year tabs)":
        url_input = st.text_input(
            "Google Sheets URL (CSV or TSV)",
            value=default_url,
            placeholder="Paste Google Sheet publish link (supports ?output=csv or ?output=tsv)...",
            label_visibility="collapsed"
        )
    elif input_mode == "Single URL (One specific tab)":
        st.caption("Loads only the tab specified in the URL (e.g., 'Todays Sales', 'Last Day Sales')")
        url_input = st.text_input(
            "Tab URL (with ?gid=XXX if needed)",
            value="https://docs.google.com/spreadsheets/d/e/2PACX-1vTOiRkybNzMNvEaLxSFsX0nGIiM07BbNVsBbsX1dG8AmGOmSu8baPrVYL0cOqoYN4tRWUj1UjUbH1Ij/pub?output=csv",
            placeholder="Paste specific tab URL...",
            label_visibility="collapsed"
        )
        single_tab_mode = True
    elif input_mode == "Multiple URLs (One per year)":
        # Manual URL input for each year
        st.caption("Enter individual URLs for each year tab (leave blank to skip):")
        url_cols = st.columns(3)
        current_year = datetime.now().year
        years = list(range(2021, current_year + 1))
        
        for i, year in enumerate(years):
            with url_cols[i % 3]:
                year_url = st.text_input(
                    f"{year} URL",
                    value="",
                    placeholder=f"https://.../pub?gid=...&output=csv",
                    key=f"ce_url_{year}"
                )
                if year_url.strip():
                    year_urls[str(year)] = year_url.strip()
    else:  # Upload File(s)
        st.caption("Upload CSV, TSV, or Excel files. Filename should contain year (e.g., 'orders_2025.csv'):")
        files = st.file_uploader(
            "Upload files",
            type=["csv", "tsv", "txt", "xlsx", "xls"],
            accept_multiple_files=True,
            key="ce_file_upload"
        )
        if files:
            for file in files:
                # Try to detect year from filename
                year_match = re.search(r'\b(20\d{2})\b', file.name)
                year = year_match.group(1) if year_match else "Unknown"
                uploaded_files[year] = file
    
    registry_file = st.text_input("Registry file (Excel)", value="customer_registry.xlsx")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("🔄 Extract & Process Customers", type="primary", use_container_width=True):
            with st.spinner("Processing..."):
                try:
                    if year_urls:
                        # Use manual year URLs
                        customers_df, metadata, raw_df = extract_customers_from_year_urls(
                            year_urls=year_urls,
                            save_registry_path=registry_file
                        )
                    elif uploaded_files:
                        # Use uploaded files
                        customers_df, metadata, raw_df = extract_customers_from_uploaded_files(
                            uploaded_files=uploaded_files,
                            save_registry_path=registry_file
                        )
                    elif single_tab_mode:
                        # Use single tab mode - loads just one tab without year discovery
                        customers_df, metadata, raw_df = extract_customers_from_single_tab(
                            url=url_input.strip(),
                            save_registry_path=registry_file
                        )
                    else:
                        # Use single URL auto-discovery (year tabs)
                        customers_df, metadata, raw_df = extract_customers_from_google_sheet(
                            url=url_input.strip(),
                            save_registry_path=registry_file
                        )
                    st.session_state[_SESSION_KEY] = customers_df
                    st.session_state["ce_metadata"] = metadata
                    st.session_state["ce_raw"] = raw_df
                    st.rerun()
                except MemoryError as e:
                    gc.collect()
                    st.error("⚠️ Memory Error: The system ran out of memory while processing.")
                    st.warning(f"Details: {str(e)}")
                    st.info("""
                    💡 **Recovery Options:**
                    1. **Close other applications** to free up RAM
                    2. **Process fewer years** - try loading one year at a time
                    3. **Restart the Streamlit app** to clear memory
                    4. **Use the 'Single URL (One specific tab)' mode** for smaller datasets
                    5. **Split your data** - divide the Google Sheet into smaller chunks
                    
                    The app is still running. You can try again with a different approach.
                    """)
                except Exception as e:
                    st.error(f"Error: {str(e)}")
    
    with col2:
        if st.button("📂 Load Existing Registry", use_container_width=True):
            registry = load_registry(registry_file)
            if registry is not None:
                st.session_state[_SESSION_KEY] = registry
                st.success(f"✅ Loaded {len(registry):,} customers from registry")
                st.rerun()
            else:
                st.warning("No existing registry found")
    
    # ── Display Results ───────────────────────────────────────────────────
    customers_df = st.session_state.get(_SESSION_KEY)
    metadata = st.session_state.get("ce_metadata")
    
    if customers_df is None:
        st.info("👆 Click 'Extract & Process Customers' to begin.")
        return
    
    # KPI Cards
    st.markdown("---")
    st.markdown("#### 📈 Summary")
    
    k1, k2, k3, k4 = st.columns(4)
    _metric_card(k1, "Unique Customers", f"{len(customers_df):,}", "👥")
    _metric_card(k2, "With Phone", f"{(customers_df['primary_phone'] != '').sum():,}", "📱")
    _metric_card(k3, "With Email", f"{(customers_df['primary_email'] != '').sum():,}", "📧")
    _metric_card(k4, "With Both", 
                 f"{((customers_df['primary_phone'] != '') & (customers_df['primary_email'] != '')).sum():,}", "✅")
    
    if metadata:
        st.caption(f"New customers added: **{metadata.get('new_customers_added', 0):,}** | "
                   f"Last run: {metadata.get('last_run', 'N/A')[:19]}")
    
    # ── Customer Table ────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 🧑‍💼 Unique Customer Registry")
    
    # Search and filters
    sf1, sf2, sf3 = st.columns([2, 2, 2])
    search = sf1.text_input("🔍 Search by name, phone, or email", key="ce_search")
    
    # Date range filter
    from datetime import date, timedelta
    
    # Parse date columns
    date_min_col = "first_order_date"
    date_max_col = "last_order_date"
    
    # Get overall date range from data
    if date_min_col in customers_df.columns:
        df_dates = pd.to_datetime(customers_df[date_min_col], errors='coerce')
        overall_min = df_dates.min().date() if df_dates.notna().any() else date(2021, 1, 1)
        overall_max = date.today()
    else:
        overall_min, overall_max = date(2021, 1, 1), date.today()
    
    # Quick range selector
    quick_range = sf2.selectbox(
        "📅 Quick Date Range",
        ["All Time", "2021", "2022", "2023", "2024", "2025", "2026", 
         "Last 30 days", "Last 90 days", "Last 6 months", "Last 12 months"],
        key="ce_quick_range"
    )
    
    today = date.today()
    if quick_range == "All Time":
        start_d, end_d = date(2021, 1, 1), today
    elif quick_range == "2021":
        start_d, end_d = date(2021, 1, 1), date(2021, 12, 31)
    elif quick_range == "2022":
        start_d, end_d = date(2022, 1, 1), date(2022, 12, 31)
    elif quick_range == "2023":
        start_d, end_d = date(2023, 1, 1), date(2023, 12, 31)
    elif quick_range == "2024":
        start_d, end_d = date(2024, 1, 1), date(2024, 12, 31)
    elif quick_range == "2025":
        start_d, end_d = date(2025, 1, 1), date(2025, 12, 31)
    elif quick_range == "2026":
        start_d, end_d = date(2026, 1, 1), date(2026, 12, 31)
    elif quick_range == "Last 30 days":
        start_d, end_d = today - timedelta(days=30), today
    elif quick_range == "Last 90 days":
        start_d, end_d = today - timedelta(days=90), today
    elif quick_range == "Last 6 months":
        start_d, end_d = today - timedelta(days=180), today
    elif quick_range == "Last 12 months":
        start_d, end_d = today - timedelta(days=365), today
    else:
        start_d, end_d = overall_min, overall_max
    
    # Custom date range override
    with sf3.expander("📆 Custom Range"):
        custom_start = st.date_input("From", value=start_d, min_value=date(2021, 1, 1), max_value=today, key="ce_start")
        custom_end = st.date_input("To", value=end_d, min_value=date(2021, 1, 1), max_value=today, key="ce_end")
    
    # Use custom dates if they differ from quick range
    if custom_start != start_d or custom_end != end_d:
        start_d, end_d = custom_start, custom_end
    
    # Apply filters
    report_view = customers_df.copy()
    
    # Ensure string type for search columns
    str_cols = ["primary_name", "primary_phone", "primary_email", "all_names"]
    for col in str_cols:
        if col in report_view.columns:
            report_view[col] = report_view[col].fillna('').astype(str)
    
    # Text search filter
    if search.strip():
        mask = (
            report_view["primary_name"].str.contains(search, case=False, na=False)
            | report_view["primary_phone"].str.contains(search, case=False, na=False)
            | report_view["primary_email"].str.contains(search, case=False, na=False)
            | report_view["all_names"].str.contains(search, case=False, na=False)
        )
        report_view = report_view[mask]
    
    # Date range filter - show customers whose first order is within range
    # OR whose last order is within range (active in period)
    if date_min_col in report_view.columns:
        report_view[date_min_col] = pd.to_datetime(report_view[date_min_col], errors='coerce')
        report_view[date_max_col] = pd.to_datetime(report_view[date_max_col], errors='coerce')
        
        # Convert Python date objects to pandas Timestamp for comparison
        start_ts = pd.Timestamp(start_d)
        end_ts = pd.Timestamp(end_d)
        
        date_mask = (
            (report_view[date_min_col] >= start_ts) & (report_view[date_min_col] <= end_ts)
        ) | (
            (report_view[date_max_col] >= start_ts) & (report_view[date_max_col] <= end_ts)
        ) | (
            (report_view[date_min_col] <= start_ts) & (report_view[date_max_col] >= end_ts)
        )  # Active throughout period
        report_view = report_view[date_mask]
    
    st.caption(f"Showing **{len(report_view):,}** customers | Date range: **{start_d} to {end_d}**")
    
    # Display columns selection
    display_cols = st.multiselect(
        "Columns to display",
        options=customers_df.columns.tolist(),
        default=["customer_id", "primary_name", "primary_phone", "primary_email", 
                 "total_spent", "order_count", "first_order_date", "last_order_date"],
        key="ce_display_cols"
    )
    
    if display_cols:
        st.dataframe(report_view[display_cols], use_container_width=True, height=500)
    else:
        st.dataframe(report_view, use_container_width=True, height=500)
    
    # ── Downloads ──────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 💾 Downloads")
    
    d1, d2, d3 = st.columns(3)
    
    with d1:
        csv_bytes = report_view.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "⬇️ Download as CSV",
            data=csv_bytes,
            file_name="customer_registry.csv",
            mime="text/csv",
            use_container_width=True,
        )
    
    with d2:
        if st.button("📊 Export Full Excel Report", use_container_width=True):
            raw_df = st.session_state.get("ce_raw")
            report_name = f"customer_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            if export_full_report(customers_df, raw_df, report_name):
                st.success(f"Report saved: `{report_name}`")
                with open(report_name, "rb") as f:
                    st.download_button(
                        "⬇️ Download Excel Report",
                        f,
                        file_name=report_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
    
    with d3:
        excel_buffer = io.BytesIO()
        report_view.to_excel(excel_buffer, index=False, engine='openpyxl')
        excel_buffer.seek(0)
        st.download_button(
            "⬇️ Download Filtered Excel",
            data=excel_buffer,
            file_name="filtered_customers.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
