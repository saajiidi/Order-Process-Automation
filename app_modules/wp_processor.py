import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, Optional
import urllib.parse
import re
import io

class WhatsAppOrderProcessor:
    def __init__(self, config: Optional[Dict] = None):
        """Initialize with configuration for column mappings."""
        self.config = config or {
            'phone_col': 'Phone (Billing)',
            'name_col': 'Full Name (Billing)',
            'order_id_col': 'Order ID',
            'product_col': 'Product Name (main)',
            'sku_col': 'SKU',
            'quantity_col': 'Quantity',
            'price_col': 'Item cost',
            'payment_method_col': 'Payment Method Title',
            'address_col': 'Address 1&2 (Billing)',
            'order_total_col': 'Order Total Amount',
            'city_col': 'City, State, Zip (Billing)'
        }

    def clean_phone_number(self, phone: str) -> str:
        """Clean and format phone number for WhatsApp."""
        if pd.isna(phone):
            return ""
        phone = ''.join(filter(str.isdigit, str(phone)))
        if phone.startswith('0'):
            return '0' + phone[1:]
        elif not phone.startswith(('88', '+88')):
            return '0' + phone
        return phone

    def format_text(self, text: str) -> str:
        """Standardize text formatting (capitalization, spacing)."""
        if not isinstance(text, str) or not text.strip():
            return ""
        
        # Helper to capitalize words properly
        def capitalize_word(w):
            if '-' in w:
                return '-'.join(p.capitalize() for p in w.split('-'))
            if "'" in w:
                return "'".join(p.capitalize() for p in w.split("'"))
            if '.' in w and not w.endswith('.'):
                return '. '.join(p.capitalize() for p in w.split('.'))
            return w.capitalize()

        # Handle comma-separated parts (addresses)
        if ',' in text:
            parts = [p.strip() for p in text.split(',')]
            formatted_parts = []
            for part in parts:
                if not part: continue
                formatted_parts.append(' '.join(capitalize_word(w) for w in part.split()))
            return ', '.join(formatted_parts)
        
        # Handle regular text
        text = re.sub(r'\.(\w)', r'. \1', text.strip())
        return ' '.join(capitalize_word(w) for w in text.split())

    def format_name(self, name):
        return self.format_text(str(name)) if pd.notna(name) else ""

    def format_address(self, *address_parts):
        parts = [self.format_text(str(p)) for p in address_parts if pd.notna(p) and str(p).strip()]
        return '\n'.join(parts)

    def detect_gender_salutation(self, name: str) -> str:
        """Detect gender from name for appropriate salutation."""
        if not isinstance(name, str):
            return "Sir"
        
        name_lower = name.lower()
        female_indicators = {
            'ms', 'miss', 'mrs', 'mst', 'begum', 'khatun', 'akter', 'parvin', 
            'sultana', 'jahan', 'bibi', 'rani', 'devi', 'nahar', 'ferdous', 
            'ara', 'banu', 'fatema', 'aisha', 'khadija', 'nusrat', 'farhana', 
            'sadia', 'jannatul', 'sumaiya', 'tanjina', 'fariha', 'sharmin',
            'nasrin', 'salma', 'shirin', 'rumana', 'sabina', 'moumita'
        }
        
        # Check parts for indicators
        parts = name_lower.replace('.', ' ').split()
        if any(part in female_indicators for part in parts):
            return "Madam"
        return "Sir"

    def _find_best_column(self, df_cols, target_keys, default):
        """Find the best matching column from a list of keys."""
        cols_lower = [str(c).lower().strip() for c in df_cols]
        for key in target_keys:
            key_l = key.lower()
            if key_l in cols_lower:
                return df_cols[cols_lower.index(key_l)]
        # Try partial match fallback
        for key in target_keys:
            key_l = key.lower()
            for i, c_l in enumerate(cols_lower):
                if key_l in c_l:
                    return df_cols[i]
        return default

    def process_orders(self, df: pd.DataFrame) -> pd.DataFrame:
        """Process raw dataframe into grouped orders with fuzzy matching."""
        df.columns = [str(col).strip() for col in df.columns]
        
        # Fuzzy Match Configuration Columns
        mapping = {
            'phone_col': (['phone', 'mobile', 'contact', 'billing phone'], 'Phone (Billing)'),
            'name_col': (['full name', 'billing name', 'name', 'first name', 'customer'], 'Full Name (Billing)'),
            'order_id_col': (['order id', 'order #', 'id', 'order number'], 'Order ID'),
            'product_col': (['product name', 'item name', 'product', 'item'], 'Product Name (main)'),
            'sku_col': (['sku', 'item sku', 'code'], 'SKU'),
            'quantity_col': (['quantity', 'qty', 'item qty'], 'Quantity'),
            'price_col': (['item cost', 'price', 'cost'], 'Item cost'),
            'payment_method_col': (['payment method', 'gateway'], 'Payment Method Title'),
            'address_col': (['address', 'shipping address', 'billing address'], 'Address 1&2 (Billing)'),
            'order_total_col': (['order total', 'total amount', 'total'], 'Order Total Amount'),
            'city_col': (['city', 'state', 'zip', 'location'], 'City, State, Zip (Billing)')
        }

        # Update config with whatever we find in the dataframe
        for config_key, (targets, default) in mapping.items():
            best_col = self._find_best_column(df.columns, targets, default)
            self.config[config_key] = best_col

        # Validate core required columns are present
        required = ['phone_col', 'name_col', 'product_col']
        missing = [self.config[col] for col in required if self.config[col] not in df.columns]
        if missing:
            raise ValueError(f"Required data columns not found. Partial match failed for: {missing}")

        phone_col = self.config['phone_col']
        
        # Clean data
        df[phone_col] = df[phone_col].apply(self.clean_phone_number)
        df[self.config['name_col']] = df[self.config['name_col']].apply(self.format_name)
        
        # Format address columns
        for col_type in ['address_col', 'city_col']:
            col_name = self.config.get(col_type)
            if col_name and col_name in df.columns:
                 df[col_name] = df[col_name].apply(lambda x: self.format_text(str(x)) if pd.notna(x) else "")

        # SKU Integration
        product_col = self.config['product_col']
        sku_col = self.config.get('sku_col')
        if sku_col and sku_col in df.columns:
            def combine(row):
                p = str(row[product_col]) if pd.notna(row[product_col]) else ""
                s = str(row[sku_col]) if pd.notna(row[sku_col]) else ""
                return f"{p} - {s}" if s.strip() and s.lower() != 'nan' else p
            df[product_col] = df.apply(combine, axis=1)

        # Aggregation rules
        agg_funcs = {
            self.config['name_col']: 'first',
            self.config['product_col']: lambda x: '\n- '.join(x),
        }

        order_id_col = self.config.get('order_id_col')
        quantity_col = self.config.get('quantity_col')
        price_col = self.config.get('price_col')

        if order_id_col in df.columns:
            agg_funcs[order_id_col] = lambda x: ', '.join(map(str, x.unique()))
        if quantity_col in df.columns:
            agg_funcs[quantity_col] = lambda x: '\n- '.join(map(str, x))
        if price_col in df.columns:
            agg_funcs[price_col] = lambda x: '\n- '.join(map(str, x))
        
        # Add optional columns to aggregation
        for key in ['address_col', 'city_col', 'payment_method_col']:
            col = self.config.get(key)
            if col and col in df.columns:
                agg_funcs[col] = 'first'

        grouped_df = df.groupby(phone_col, as_index=False).agg(agg_funcs)

        for col in [order_id_col, quantity_col, price_col]:
            if col and col not in grouped_df.columns:
                grouped_df[col] = ""

        # Calculate correct total amount (sum of unique order totals per phone)
        # This prevents summing the "Order Total" multiple times for multi-item orders
        total_col = self.config['order_total_col']
        if total_col in df.columns:
            # Get one row per order to capture the Order Total once
            if order_id_col in df.columns:
                unique_orders = df[[phone_col, order_id_col, total_col]].drop_duplicates(
                    subset=[order_id_col]
                )
            else:
                unique_orders = df[[phone_col, total_col]].drop_duplicates()

            phone_totals = unique_orders.groupby(phone_col)[total_col].sum()
            grouped_df[total_col] = grouped_df[phone_col].map(phone_totals).fillna(0)
        elif price_col in df.columns and quantity_col in df.columns:
            price_num = pd.to_numeric(df[price_col], errors="coerce").fillna(0)
            qty_num = pd.to_numeric(df[quantity_col], errors="coerce").fillna(0)
            line_total = price_num * qty_num
            phone_totals = line_total.groupby(df[phone_col]).sum()
            grouped_df[total_col] = grouped_df[phone_col].map(phone_totals).fillna(0)
        else:
            grouped_df[total_col] = 0

        return grouped_df

    def create_whatsapp_links(
        self, 
        df: pd.DataFrame, 
        custom_intro: str = None, 
        custom_footer: str = None,
        progress_callback: Optional[callable] = None,
        shorten_urls: bool = False
    ) -> pd.DataFrame:
        """Generate formatted WhatsApp messages and links."""
        df['whatsapp_link'] = None
        phone_col = self.config['phone_col']
        total_rows = len(df)

        import pyshorteners
        shortener = pyshorteners.Shortener() if shorten_urls else None
        
        for idx, row in df.iterrows():
            if progress_callback:
                progress_callback(idx + 1, total_rows)
                
            phone = row[phone_col]
            if not phone: continue

            # Determine Salutation
            name = row[self.config['name_col']]
            salutation = self.detect_gender_salutation(name)
            
            # Format Address
            formatted_address = self.format_address(
                row.get(self.config['address_col'], ''),
                row.get(self.config.get('city_col'), '')
            )

            # Build Message Intro
            order_id_val = row.get(self.config.get('order_id_col'), "")
            if custom_intro:
                intro_text = custom_intro.replace("{name}", str(name)).replace("{salutation}", salutation).replace("{order_id}", str(order_id_val))
                lines = [intro_text]
                lines.extend(["", "*Your Order:*"])
            else:
                lines = [
                    f"*Order Verification From DEEN Commerce*",
                    "",
                    f"Assalamu Alaikum, {salutation}!",
                    "",
                    f"Dear {name},",
                    "",
                    "Please verify your order details:",
                    "",
                    f"*Order ID:* {order_id_val}",
                    "",
                    "*Your Order:*",
                ]

            # Products
            products = str(row.get(self.config.get('product_col'), "")).split('\n- ')
            quantities = []
            if self.config.get('quantity_col'):
                quantities = str(row.get(self.config.get('quantity_col'), "")).split('\n- ')
            prices = []
            if self.config.get('price_col'):
                prices = str(row.get(self.config.get('price_col'), "")).split('\n- ')
            
            for i, prod in enumerate(products):
                item_line = f"- {prod.strip()}"
                if i < len(quantities): item_line += f" - Qty: {quantities[i].strip()}"
                if i < len(prices): item_line += f" - Price: {prices[i].strip()} BDT"
                lines.append(item_line)

            # Totals & Payment Logic
            total_col = self.config.get('order_total_col')
            raw_total = row.get(total_col, 0)
            total_amount = pd.to_numeric(raw_total, errors="coerce")
            total_amount = 0.0 if pd.isna(total_amount) else float(total_amount)
            collectable_amount = total_amount
            
            payment_col = self.config.get('payment_method_col')
            if payment_col and payment_col in row and pd.notna(row[payment_col]):
                method = str(row[payment_col]).lower()
                if any(x in method for x in ['bkash', 'online', 'ssl', 'paid']):
                    collectable_amount = 0

            lines.append("")
            if collectable_amount == 0:
                lines.append(f"*Total Amount:* {total_amount:.2f} BDT (PAID)")
                lines.append(f"*Collectable Amount:* 0 BDT")
            else:
                lines.append(f"*Total Amount:* {total_amount:.2f} BDT")
            
            lines.extend([
                "",
                "*Shipping Address:*",
                formatted_address,
                "",
            ])
            
            if custom_footer:
                lines.extend(custom_footer.split("\n"))
            else:
                lines.extend([
                    "Please confirm the order and address.",
                    "If any correction is needed, please let us know the possible adjustment.",
                    "",
                    "*Delivery fees apply for returns.*",
                    "",
                    "Thank you for shopping with DEEN Commerce! Grab our latest collection on: https://deencommerce.com/"
                ])

            message = "\n".join(lines)
            encoded_message = urllib.parse.quote(message)
            full_link = f"https://wa.me/+88{phone}?text={encoded_message}"
            
            if shorten_urls:
                try:
                    full_link = shortener.tinyurl.short(full_link)
                except:
                    pass # Fallback to long link if shortening fails
                    
            df.at[idx, 'whatsapp_link'] = full_link
        
        return df

    def generate_excel_bytes(self, df: pd.DataFrame) -> bytes:
        """Create formatted Excel file in memory."""
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Orders')
            worksheet = writer.sheets['Orders']
            
            # Styles
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            link_font = Font(color='0000FF', underline='single')
            center_align = Alignment(horizontal='center')
            
            # Header Formatting
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            
            # Link Formatting
            if 'whatsapp_link' in df.columns:
                whatsapp_col = df.columns.get_loc('whatsapp_link') + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=whatsapp_col)
                    if cell.value:
                        cell.hyperlink = cell.value
                        cell.value = 'Send WhatsApp'
                        cell.font = link_font
                        cell.alignment = center_align

            # Auto-width
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                worksheet.column_dimensions[column_letter].width = min((max_length + 2) * 1.2, 50)
                
        return output.getvalue()
